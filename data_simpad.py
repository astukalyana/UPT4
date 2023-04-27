from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

#MENDAPATKAN DATA WP DARI EXCEL
wb = load_workbook("./excel/wpairtanah.xlsx")
ws = wb.active

daftar_wp = []

for i in ws['A']:
    if(i.row == 1):
        continue

    j = i.value

    formatted = j[:1] + "." + j[1:2] + "." + j[2:9] + "." + j[9:11] + "." + j[11:13]
    daftar_wp.append(formatted)

print(daftar_wp)

#MEMBUAT EXCEL BARU
wb = Workbook()
ws = wb.active

#SELENIUM SIMPAD
driver = webdriver.Chrome()

driver.get("https://pajak.tangerangkab.go.id/login")

dismiss_btn = driver.find_element(By.XPATH, "//*[@id='myModal1']/div/div/div[3]/button")
dismiss_btn.click()

username = driver.find_element(By.NAME, "userid")
username.send_keys("UPTWIL4")

password = driver.find_element(By.NAME, "passwd")
password.send_keys("TOMI")
password.send_keys(Keys.RETURN)

#LOOP THROUGH DAFTAR_WP LIST
for i in daftar_wp[:1]:
    driver.get("https://pajak.tangerangkab.go.id/pad_tangerangkab/objek_pajak/")

    #TEMPORARY LIST TO APPEND TO EXCEL
    row = []

    #GET SEARCH TEXT BOX AND SEARCH THE CORRESPONDING WP
    search_bar = driver.find_element(By.XPATH, "//*[@id='table1_filter']/label/input")
    search_bar.send_keys(i)
    search_bar.send_keys(Keys.RETURN)

    input("TUNGGU")
    # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, "//*[@id='table1']/tbody/tr/td[5]")))

    driver.find_element(By.XPATH, "//*[@id='table1']/tbody/tr/td[5]").click()
    driver.find_element(By.XPATH, "//*[@id='btn_edit']").click()

    npwpd = driver.find_element(By.XPATH, "//*[@id='npwpd']").get_attribute('value')
    row.append(npwpd)

    nama_wp = driver.find_element(By.XPATH, "//*[@id='opnm']").get_attribute('value')
    row.append(nama_wp)

    kecamatan = Select(driver.find_element(By.XPATH, "//*[@id='kecamatan_id']")).first_selected_option.text
    row.append(kecamatan)

    kelurahan = Select(driver.find_element(By.XPATH, '//*[@id="kelurahan_id"]')).first_selected_option.text
    row.append(kelurahan)

    cara_hitung = Select(driver.find_element(By.XPATH, '//*[@id="air_isflat"]')).first_selected_option.text
    row.append(cara_hitung)

    zona = Select(driver.find_element(By.XPATH, '//*[@id="air_zona_id"]')).first_selected_option.text
    row.append(zona)

    manfaat = Select(driver.find_element(By.XPATH, '//*[@id="air_manfaat_id"]')).first_selected_option.text
    row.append(manfaat)

    alamat = driver.find_element(By.XPATH, '//*[@id="opalamat"]').get_attribute('value')
    row.append(alamat)
    
    print(row)
    ws.append(row)

wb.save("datasimpad.xlsx")
input("ENTER")
