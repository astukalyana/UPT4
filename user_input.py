from openpyxl import load_workbook
import pyinputplus as pyip

#MENDAPATKAN DATA WP DARI EXCEL
wb = load_workbook("./excel/datasimpad.xlsx")
ws = wb.active

daftar_wp = {}

for i in ws['A']:
    nama_wp = i.offset(column = 1).value
    daftar_wp[nama_wp] = {}
    daftar_wp[nama_wp]["nama"] = nama_wp
    daftar_wp[nama_wp]["npwpd"] = i.value
    daftar_wp[nama_wp]["metode"] = i.offset(column = 4).value
    daftar_wp[nama_wp]["zona"] = i.offset(column = 5).value
    daftar_wp[nama_wp]["Manfaat"] = i.offset(column = 6).value

print(daftar_wp)

#INPUT DATA
def get_data():
    wajib_pajak = pyip.inputMenu([i for i in daftar_wp], numbered = True)
    npwpd = daftar_wp[wajib_pajak]
    print(wajib_pajak, npwpd, sep = " - ")

    return daftar_wp[wajib_pajak]

def get_masa():
    seleksi_mode = pyip.inputMenu(["Bulan", "Bulan berurutan", "Bulan tidak berurut"], numbered = True)
    list_bulan = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    masa_pajak = {}

    banyak_tahun = int(input("Berapa banyak tahun?  "))
    for i in range(banyak_tahun):
        tahun_pajak = input("Tahun pajak:    ")
        masa_pajak[tahun_pajak] = []

        #INPUT PROCESSING
        if(seleksi_mode == "Bulan"):
            masa_pajak.append(pyip.inputMenu(list_bulan, numbered = True, prompt = "Masa pajak: \n"))
            print(masa_pajak)
        elif(seleksi_mode == "Bulan berurutan"):
            masa_awal = list_bulan.index(pyip.inputMenu(list_bulan, numbered = True, prompt = "Masa pajak awal: \n"))
            masa_akhir = list_bulan.index(pyip.inputMenu(list_bulan, numbered = True, prompt = "Masa pajak akhir: \n"))

            for i in range(masa_awal, masa_akhir + 1):
                masa_pajak[tahun_pajak].append(list_bulan[i])
            
        else:
            banyak_bulan = int(input("Berapa banyak bulan? "))

            for i in range(banyak_bulan):
                masa_pajak[tahun_pajak].append(pyip.inputMenu(list_bulan, numbered = True, prompt = "Masa pajak: \n"))
            
    print(masa_pajak)
    return masa_pajak  

wp = get_data()
print(wp)